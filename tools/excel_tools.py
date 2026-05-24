"""M.A.Y.D.A.Y Excel Tools — openpyxl read/write"""
import logging
from tools.base_tool import BaseTool
logger = logging.getLogger('mayday.tools.excel')

class ExcelTools(BaseTool):
    @property
    def name(self) -> str: return 'excel'
    @property
    def description(self) -> str: return 'Excel read/write operations'
    def get_capabilities(self) -> list[str]:
        return ['excel_read', 'excel_write', 'excel_create']

    def execute(self, parameters: dict) -> dict:
        name = parameters.get('_tool_name', '')
        if 'write' in name: return self._write(parameters)
        if 'create' in name: return self._create(parameters)
        return self._read(parameters)

    def _read(self, params: dict) -> dict:
        import openpyxl
        path = params.get('path', '')
        sheet = params.get('sheet', None)
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb[sheet] if sheet else wb.active
            data = [[cell.value for cell in row] for row in ws.iter_rows()]
            wb.close()
            return {'status': 'success', 'data': data, 'rows': len(data)}
        except Exception as e:
            return {'status': 'error', 'error': str(e)}

    def _write(self, params: dict) -> dict:
        import openpyxl
        path = params.get('path', '')
        data = params.get('data', [])
        sheet = params.get('sheet', 'Sheet1')
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
            for r, row in enumerate(data, 1):
                for c, val in enumerate(row, 1):
                    ws.cell(row=r, column=c, value=val)
            wb.save(path); wb.close()
            return {'status': 'success', 'rows_written': len(data)}
        except Exception as e:
            return {'status': 'error', 'error': str(e)}

    def _create(self, params: dict) -> dict:
        import openpyxl
        from pathlib import Path
        path = params.get('path', 'output.xlsx')
        data = params.get('data', [])
        try:
            Path(path).parent.mkdir(parents=True, exist_ok=True)
            wb = openpyxl.Workbook()
            ws = wb.active
            for r, row in enumerate(data, 1):
                for c, val in enumerate(row, 1):
                    ws.cell(row=r, column=c, value=val)
            wb.save(path); wb.close()
            file_path = Path(path)
            return {'status': 'success', 'path': str(file_path.resolve()), 'bytes': file_path.stat().st_size, 'exists': file_path.exists()}
        except Exception as e:
            return {'status': 'error', 'error': str(e)}
